"""
app.py  —  Physiotherapy Lead Extractor  (Streamlit UI)
Tabs:
  1. Single Text entry
  2. Bulk Paste
  3. Upload File
  4. Gap Analysis + ML Data Prep  ← NEW
"""

import streamlit as st
import pandas as pd
import io
import re
import json
from datetime import datetime
from collections import Counter
from physio_ner import extract_physio_leads, PHYSIO_KEYWORDS

# ── Page config ───────────────────────────────────────────────────────────────
st.set_page_config(page_title="Physio Lead Extractor", page_icon="🏥", layout="wide")

# ── CSS ───────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=DM+Sans:wght@400;500;600;700&family=DM+Mono:wght@400;500&display=swap');
html, body, [class*="css"] { font-family: 'DM Sans', sans-serif; }
.stApp { background: #f0f4f8; }

[data-testid="stSidebar"] { background-color: #1a365d !important; }
[data-testid="stSidebar"] * { color: #ffffff !important; }
[data-testid="stSidebar"] .stSelectbox > div > div,
[data-testid="stSidebar"] .stMultiSelect > div > div,
[data-testid="stSidebar"] input,
[data-testid="stSidebar"] textarea { background-color: #243f6a !important; border-color: #3a5a8a !important; }
[data-testid="stSidebar"] [data-baseweb="tag"] { background-color: #3a5a8a !important; }
[data-testid="stSidebar"] hr { border-color: #3a5a8a !important; }

.main-header {
    background: linear-gradient(135deg, #1a365d 0%, #2d6a9f 100%);
    color: white; padding: 1.8rem 2.5rem; border-radius: 16px;
    margin-bottom: 1.5rem; box-shadow: 0 4px 20px rgba(26,54,93,0.3);
}
.main-header h1 { margin:0; font-size:1.8rem; font-weight:700; }
.main-header p  { margin:0.3rem 0 0; opacity:0.8; font-size:0.9rem; }

.card { background:white; border-radius:14px; padding:1.5rem;
        box-shadow:0 2px 12px rgba(0,0,0,0.07); margin-bottom:1rem; }

.stat-box { background:white; border-radius:12px; padding:1.2rem 1rem;
            text-align:center; box-shadow:0 2px 10px rgba(0,0,0,0.06); }
.stat-num   { font-size:2rem; font-weight:700; color:#1a365d; }
.stat-label { font-size:0.78rem; color:#6b7280; margin-top:2px; }

.result-physio { border-left:4px solid #16a34a; background:#f0fdf4;
                 padding:1rem 1.2rem; border-radius:0 10px 10px 0; margin-bottom:0.8rem; }
.result-none   { border-left:4px solid #94a3b8; background:#f8fafc;
                 padding:1rem 1.2rem; border-radius:0 10px 10px 0; margin-bottom:0.8rem; }
.result-review { border-left:4px solid #f59e0b; background:#fffbeb;
                 padding:1rem 1.2rem; border-radius:0 10px 10px 0; margin-bottom:0.8rem; }

.label-physio { background:#16a34a; color:white; padding:3px 12px;
                border-radius:20px; font-size:0.78rem; font-weight:600; }
.label-none   { background:#94a3b8; color:white; padding:3px 12px;
                border-radius:20px; font-size:0.78rem; font-weight:600; }
.label-review { background:#f59e0b; color:white; padding:3px 12px;
                border-radius:20px; font-size:0.78rem; font-weight:600; }

.kw-chip { display:inline-block; background:#dbeafe; color:#1e40af;
           padding:2px 10px; border-radius:12px; font-size:0.75rem;
           font-weight:500; margin:2px; }

.conf-bar-wrap { background:#e5e7eb; border-radius:6px; height:8px; width:100%; margin-top:4px; }
.conf-bar      { border-radius:6px; height:8px; }

.gap-stat { background:white; border-radius:12px; padding:1rem;
            text-align:center; box-shadow:0 2px 8px rgba(0,0,0,0.06); }
.gap-num   { font-size:1.8rem; font-weight:700; }
.gap-label { font-size:0.75rem; color:#6b7280; margin-top:2px; }

.step-box { background:white; border-radius:12px; padding:1.2rem 1.5rem;
            border-left:4px solid #2d6a9f; margin-bottom:0.8rem;
            box-shadow:0 2px 8px rgba(0,0,0,0.05); }
.step-num { font-size:0.72rem; font-weight:700; color:#2d6a9f;
            text-transform:uppercase; letter-spacing:1px; }

.stTabs [data-baseweb="tab-list"] { background:white; border-radius:12px;
    padding:6px; box-shadow:0 2px 8px rgba(0,0,0,0.06); margin-bottom:1rem; gap:6px; }
.stTabs [data-baseweb="tab"] { border-radius:8px !important; font-weight:500; padding:8px 18px; }
.stTabs [aria-selected="true"] { background-color:#1a365d !important; color:white !important; }
</style>
""", unsafe_allow_html=True)

# ── Header ────────────────────────────────────────────────────────────────────
st.markdown("""
<div class="main-header">
    <h1>🏥 Physiotherapy Lead Extractor</h1>
    <p>NER-powered · Rule-Based Phase 1 · Extracts physio leads from doctor discharge notes</p>
</div>
""", unsafe_allow_html=True)

# ── Sidebar ───────────────────────────────────────────────────────────────────
with st.sidebar:
    st.markdown("## ⚙️ Settings")
    st.markdown("---")
    st.markdown("### 🎯 Confidence Threshold")
    conf_threshold = st.slider(
        "Min confidence to flag as Physio",
        min_value=0.50, max_value=1.00, value=0.75, step=0.05, format="%.0%%"
    )
    st.markdown("### 📋 Keyword Dictionary")
    st.markdown(f"**{len(PHYSIO_KEYWORDS)} keywords** loaded")
    with st.expander("View all keywords"):
        for kw in sorted(PHYSIO_KEYWORDS):
            st.markdown(f"• {kw}")
    st.markdown("### ➕ Custom Keywords")
    custom_kws = st.text_area("One keyword per line:",
                              placeholder="BALANCE BOARD\nPROPRIOCEPTION", height=80)
    st.markdown("---")
    st.markdown("### 📊 Display Options")
    show_details   = st.checkbox("Show match details", value=True)
    show_no_advice = st.checkbox("Show 'No Advice' rows", value=True)

if custom_kws.strip():
    from physio_ner import matcher, nlp
    new_kws      = [k.strip().upper() for k in custom_kws.strip().split("\n") if k.strip()]
    new_patterns = [nlp.make_doc(k.lower()) for k in new_kws]
    matcher.add("PHYSIO", new_patterns)
    PHYSIO_KEYWORDS.extend(new_kws)


# ── Shared helpers ────────────────────────────────────────────────────────────
def conf_color(conf):
    if conf >= 0.90: return "#16a34a"
    if conf >= 0.75: return "#f59e0b"
    return "#94a3b8"

def detect_columns(df):
    id_cands  = ["patient id","patient_id","patientid","id","mrn","pid","patient number"]
    txt_cands = ["discharge summary","summary","text","notes","discharge_summary",
                 "clinical notes","advice","report","doctor advice","doctor notes"]
    id_col = txt_col = None
    for col in df.columns:
        cl = col.strip().lower()
        if cl in id_cands  and id_col  is None: id_col  = col
        if cl in txt_cands and txt_col is None: txt_col = col
    return id_col, txt_col

def render_result_card(pid, result):
    label = result["label"]
    conf  = result["confidence"]
    kws   = result["keywords"]

    if label == "No Advice" and not show_no_advice:
        return

    css   = "result-physio" if label == "Physiotherapy" else "result-none"
    lcss  = "label-physio"  if label == "Physiotherapy" else "label-none"
    color = conf_color(conf)
    bar_w = int(conf * 100)

    kw_html = "".join(f'<span class="kw-chip">{k}</span>' for k in kws) if kws else "<em style='color:#94a3b8'>—</em>"

    st.markdown(f"""
    <div class="{css}">
        <div style="display:flex;justify-content:space-between;align-items:center">
            <div><span style="font-weight:600">{pid}</span>&nbsp;&nbsp;
                 <span class="{lcss}">{label}</span></div>
            <div style="font-size:0.82rem">Confidence: <strong style="color:{color}">{conf:.0%}</strong></div>
        </div>
        <div class="conf-bar-wrap"><div class="conf-bar" style="width:{bar_w}%;background:{color}"></div></div>
        <div style="margin-top:8px;font-size:0.82rem"><strong>Keywords:</strong><br>{kw_html}</div>
    </div>""", unsafe_allow_html=True)

def run_ner_on_df(df, id_col, txt_col):
    results = []
    prog    = st.progress(0, text="Starting…")
    total   = len(df)
    for i, row in df.iterrows():
        pid  = str(row[id_col])
        text = str(row[txt_col]) if pd.notna(row[txt_col]) else ""
        res  = extract_physio_leads(text)
        if res["confidence"] < conf_threshold and res["label"] == "Physiotherapy":
            res["label"] = "No Advice"
        results.append((pid, text, res))
        prog.progress((i+1)/total, text=f"Processing {i+1}/{total}…")
    prog.progress(1.0, text="✅ Done!")
    return results

def render_summary_stats(results_list):
    total  = len(results_list)
    physio = sum(1 for _, _, r in results_list if r["label"] == "Physiotherapy")
    none   = total - physio

    s1, s2, s3 = st.columns(3)
    for col, num, lbl in [(s1, total, "Total Records"),
                          (s2, physio, "Physio Leads"),
                          (s3, none,   "No Advice")]:
        col.markdown(f'<div class="stat-box"><div class="stat-num">{num}</div>'
                     f'<div class="stat-label">{lbl}</div></div>', unsafe_allow_html=True)

def export_results(results_list):
    rows = []
    for pid, text, res in results_list:
        rows.append({
            "Patient ID":     pid,
            "Label":          res["label"],
            "Confidence":     res["confidence"],
            "Keywords Found": " | ".join(res["keywords"]),
            "Keyword Count":  len(res["keywords"]),
            "Raw Text":       text[:300],
        })
    return pd.DataFrame(rows)

def download_buttons(export_df, prefix="leads"):
    ts = datetime.now().strftime("%Y%m%d_%H%M")
    c1, c2 = st.columns(2)
    with c1:
        st.download_button("⬇️ Download CSV", export_df.to_csv(index=False).encode(),
                           f"{prefix}_{ts}.csv", "text/csv", use_container_width=True)
    with c2:
        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine="openpyxl") as w:
            export_df.to_excel(w, index=False, sheet_name="Results")
        st.download_button("⬇️ Download Excel", buf.getvalue(),
                           f"{prefix}_{ts}.xlsx",
                           "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                           use_container_width=True)


# ── TABS ──────────────────────────────────────────────────────────────────────
tab1, tab2, tab3, tab4 = st.tabs([
    "💬  Single Text",
    "📋  Bulk Paste",
    "📂  Upload File",
    "🔬  Gap Analysis"
])


# ════════════════════════════════════════════════════════════
# TAB 1 — SINGLE TEXT
# ════════════════════════════════════════════════════════════
with tab1:
    st.markdown("### Paste a single doctor's discharge note")
    c1, c2 = st.columns([3, 1])
    with c1:
        single_text = st.text_area("Note", height=160, label_visibility="collapsed",
            placeholder="|PHYSIOTHERAPY FOR BACK AND NECK-TO CONTINUE\n|REVIEW AFTER 1 WEEK\n|CHEST PHYSIOTHERAPY: POSTURAL DRAINAGE")
        pid_single  = st.text_input("Patient ID (optional):", value="PT001")
    with c2:
        st.markdown("""
        <div class="card">
            <strong>💡 Tips</strong><br><br>
            • Paste raw text as-is<br>
            • Pipe <code>|</code> separators OK<br>
            • Mixed case OK<br>
            • Typos handled ✓
        </div>""", unsafe_allow_html=True)

    if st.button("🔍 Extract Physio Lead", type="primary", use_container_width=True, key="btn_single"):
        if not single_text.strip():
            st.warning("Please enter some text.")
        else:
            result = extract_physio_leads(single_text)
            if result["confidence"] < conf_threshold and result["label"] == "Physiotherapy":
                result["label"] = "No Advice"
            st.markdown("---")
            st.markdown("### Result")
            render_result_card(pid_single or "Patient", result)
            if show_details and result["details"]:
                with st.expander("🔬 Match Details"):
                    st.dataframe(pd.DataFrame(result["details"]), use_container_width=True, hide_index=True)
            with st.expander("📝 Preprocessed Text"):
                st.code(result["clean_text"], language=None)


# ════════════════════════════════════════════════════════════
# TAB 2 — BULK PASTE
# ════════════════════════════════════════════════════════════
with tab2:
    st.markdown("### Paste multiple records")
    st.caption("Format: `PatientID:::Text` per line — or just paste text lines (auto-numbered)")
    bulk_text = st.text_area("Bulk input", height=220, label_visibility="collapsed",
        placeholder="PT001:::PHYSIOTHERAPY AT HOME, REVIEW AFTER 1 WEEK\nPT002:::CHEST PHYSIOTHERAPY POSTURAL DRAINAGE\nPT003:::REVIEW WITH CARDIOLOGIST, TAB ASPIRIN 75MG")

    if st.button("🔍 Extract Leads", type="primary", use_container_width=True, key="btn_bulk"):
        if not bulk_text.strip():
            st.warning("Please paste some records.")
        else:
            lines = [l.strip() for l in bulk_text.strip().split("\n") if l.strip()]
            records = []
            for i, line in enumerate(lines):
                if ":::" in line:
                    pid, text = line.split(":::", 1)
                else:
                    pid, text = f"PT{i+1:03d}", line
                records.append({"patient_id": pid.strip(), "text": text.strip()})

            results_list = []
            prog = st.progress(0)
            for j, rec in enumerate(records):
                res = extract_physio_leads(rec["text"])
                if res["confidence"] < conf_threshold and res["label"] == "Physiotherapy":
                    res["label"] = "No Advice"
                results_list.append((rec["patient_id"], rec["text"], res))
                prog.progress((j+1)/len(records))
            prog.progress(1.0)

            st.markdown("### 📊 Summary")
            render_summary_stats(results_list)
            st.markdown("### Results")
            for pid, _, res in results_list:
                render_result_card(pid, res)
            st.markdown("### 💾 Export")
            download_buttons(export_results(results_list), "bulk_leads")


# ════════════════════════════════════════════════════════════
# TAB 3 — FILE UPLOAD
# ════════════════════════════════════════════════════════════
with tab3:
    st.markdown("### Upload Excel or CSV")
    c1, c2 = st.columns([2, 1])
    with c1:
        uploaded = st.file_uploader("Upload", type=["xlsx","xls","csv"],
                                    label_visibility="collapsed")
    with c2:
        st.markdown("""
        <div class="card">
            <strong>📋 Expected Format</strong><br><br>
            • <code>Patient ID</code> column<br>
            • <code>Discharge Summary</code> column<br>
            • One row per patient<br><br>
            <span style="font-size:0.82rem;color:#6b7280">Column names auto-detected</span>
        </div>""", unsafe_allow_html=True)

    if not uploaded:
        sample = pd.DataFrame({
            "Patient ID": ["PT001","PT002","PT003"],
            "Discharge Summary": [
                "PHYSIOTHERAPY AT HOME | REVIEW AFTER 1 WEEK",
                "CHEST PHYSIOTHERAPY: POSTURAL DRAINAGE | INCENTIVE SPIROMETRY",
                "TAB PARACETAMOL 500MG BD | REVIEW WITH CARDIOLOGIST AFTER 2 WEEKS"
            ]
        })
        st.download_button("📥 Download Sample Template",
                           sample.to_csv(index=False).encode(),
                           "sample_template.csv", "text/csv")
    else:
        try:
            df = pd.read_csv(uploaded) if uploaded.name.endswith(".csv") else pd.read_excel(uploaded)
        except Exception as e:
            st.error(f"Could not read file: {e}")
            st.stop()

        st.success(f"✅ Loaded **{len(df)} rows** from `{uploaded.name}`")
        id_col, txt_col = detect_columns(df)

        c1, c2 = st.columns(2)
        with c1:
            id_col  = st.selectbox("Patient ID column:", df.columns.tolist(),
                                   index=df.columns.tolist().index(id_col) if id_col else 0)
        with c2:
            txt_col = st.selectbox("Summary column:", df.columns.tolist(),
                                   index=df.columns.tolist().index(txt_col) if txt_col else min(1, len(df.columns)-1))

        with st.expander("👁️ Preview (first 5 rows)"):
            st.dataframe(df[[id_col, txt_col]].head(5), use_container_width=True)

        if st.button("🔍 Extract Physio Leads", type="primary", use_container_width=True, key="btn_upload"):
            results_list = run_ner_on_df(df, id_col, txt_col)
            st.markdown("### 📊 Summary")
            render_summary_stats(results_list)
            st.markdown("### Results")
            for pid, _, res in results_list:
                render_result_card(pid, res)
            st.markdown("### 💾 Export")
            download_buttons(export_results(results_list), "file_leads")


# ════════════════════════════════════════════════════════════
# TAB 4 — GAP ANALYSIS
# ════════════════════════════════════════════════════════════
with tab4:
    st.markdown("### 🔬 Gap Analysis + ML Training Data Preparation")

    st.markdown("""
    <div class="step-box">
        <div class="step-num">What this does</div>
        <div style="margin-top:6px;font-size:0.92rem">
        Upload your full patient dataset → auto-label all rows → flag uncertain ones for your review →
        generate a color-coded Excel + ML training file. No command line needed.
        </div>
    </div>""", unsafe_allow_html=True)

    # ── How to use guide ──
    with st.expander("📖 How to use — read this first", expanded=True):
        st.markdown("""
**Step 1 — Upload your file below**
Your Excel/CSV must have at least:
- A **Patient ID** column (any name like `MRN`, `ID`, `Patient No`)
- A **Discharge Summary / Advice** column (all doctor advice text for that patient in one cell)

**Step 2 — Select correct columns**
The app auto-detects column names. Double-check they are correct before running.

**Step 3 — Click Run Gap Analysis**
The model processes every row and assigns:
- 🟢 **Auto-labeled Physiotherapy** — confident match (≥85%), no review needed
- 🟡 **Needs Review** — uncertain (1–84% confidence), you decide
- 🔴 **Auto-labeled No Advice** — no physio keywords found

**Step 4 — Download the Output Excel**
It has 3 sheets:
- `All Results` — color-coded full list
- `Needs Review` — only uncertain rows with empty `MANUAL_LABEL` column → **you fill this**
- `Confirmed Labels` — high-confidence rows ready for ML training

**Step 5 — Fill the Needs Review sheet**
Open the Excel, go to `Needs Review` sheet, fill `MANUAL_LABEL` column with:
`Physiotherapy` or `No Advice` for each yellow row.
If you spot a new keyword the model missed — write it in the `NOTES` column.

**Step 6 — Send back the filled Excel**
Share it with your developer — it becomes the training data for the Phase 2 ML model.
        """)

    st.markdown("---")

    # ── File upload ──
    ga_file = st.file_uploader("Upload your patient data (Excel or CSV)",
                                type=["xlsx","xls","csv"], key="gap_upload")

    if not ga_file:
        sample = pd.DataFrame({
            "Patient ID":         ["PT001","PT002","PT003","PT004","PT005"],
            "Discharge Summary":  [
                "|PHYSIOTHERAPY AT HOME|REVIEW AFTER 1 WEEK WITH DR SHARMA",
                "|CHEST PHYSIOTHERAPY: POSTURAL DRAINAGE|INCENTIVE SPIROMETRY",
                "|TAB PARACETAMOL 500MG BD|INJ CEFTRIAXONE|MONITOR VITALS",
                "|SLR EXERCISES|ANKLE PUMPS|FWB WITH WALKER|REVIEW ORTHO OPD",
                "|REVIEW WITH NEPHROLOGIST|RESTRICT FLUIDS 1L/DAY|MONITOR CREATININE",
            ]
        })
        st.download_button("📥 Download Sample Template",
                           sample.to_csv(index=False).encode(),
                           "gap_analysis_template.csv", "text/csv")
    else:
        try:
            ga_df = pd.read_csv(ga_file) if ga_file.name.endswith(".csv") else pd.read_excel(ga_file)
        except Exception as e:
            st.error(f"Could not read file: {e}")
            st.stop()

        st.success(f"✅ Loaded **{len(ga_df)} rows** from `{ga_file.name}`")

        # Column selection
        id_col_ga, txt_col_ga = detect_columns(ga_df)
        c1, c2 = st.columns(2)
        with c1:
            id_col_ga  = st.selectbox("Patient ID column:", ga_df.columns.tolist(),
                                      index=ga_df.columns.tolist().index(id_col_ga) if id_col_ga else 0,
                                      key="ga_id_col")
        with c2:
            txt_col_ga = st.selectbox("Discharge Summary column:", ga_df.columns.tolist(),
                                      index=ga_df.columns.tolist().index(txt_col_ga) if txt_col_ga else min(1, len(ga_df.columns)-1),
                                      key="ga_txt_col")

        with st.expander("👁️ Preview (first 5 rows)"):
            st.dataframe(ga_df[[id_col_ga, txt_col_ga]].head(5), use_container_width=True)

        # Thresholds
        with st.expander("⚙️ Advanced — Labeling Thresholds"):
            c1, c2 = st.columns(2)
            with c1:
                auto_physio_thresh = st.slider(
                    "Auto-label PHYSIO above this confidence",
                    0.50, 1.00, 0.85, 0.05, format="%.0%%",
                    help="Rows above this → auto-labeled Physiotherapy"
                )
            with c2:
                review_thresh = st.slider(
                    "Flag for review below this confidence",
                    0.50, 1.00, 0.85, 0.05, format="%.0%%",
                    help="Rows between 0 and this → flagged for manual review"
                )

        if st.button("🔬 Run Gap Analysis", type="primary", use_container_width=True, key="btn_gap"):

            # ── Run NER on all rows ──
            ga_results = []
            prog = st.progress(0, text="Analyzing records…")
            total = len(ga_df)

            physio_adjacent = [
                "therapy","therap","exercise","exercis","rehabilit",
                "mobiliz","mobilis","ambul","physio","strengthen",
                "stretching","gait","walker","bearing","elevation",
                "electro","magneto","spirometr","drainage","postural"
            ]

            for i, row in ga_df.iterrows():
                pid  = str(row[id_col_ga])
                text = str(row[txt_col_ga]) if pd.notna(row[txt_col_ga]) else ""
                res  = extract_physio_leads(text)
                conf = res["confidence"]

                if conf >= auto_physio_thresh:
                    auto_label   = "Physiotherapy"
                    needs_review = False
                elif conf > 0:
                    auto_label   = "NEEDS REVIEW"
                    needs_review = True
                else:
                    auto_label   = "No Advice"
                    needs_review = False

                # Check for potential missed physio hints
                clean_lower = res["clean_text"].lower()
                hints = [h for h in physio_adjacent if h in clean_lower]

                ga_results.append({
                    "patient_id":    pid,
                    "raw_text":      text,
                    "clean_text":    res["clean_text"],
                    "auto_label":    auto_label,
                    "confidence":    round(conf, 2),
                    "needs_review":  needs_review,
                    "keywords_found": " | ".join(res["keywords"]),
                    "keyword_count": len(res["keywords"]),
                    "potential_miss_hints": ", ".join(hints) if hints and auto_label == "No Advice" else "",
                })
                prog.progress((i+1)/total, text=f"Analyzing {i+1}/{total}…")

            prog.progress(1.0, text="✅ Analysis complete!")
            ga_res_df = pd.DataFrame(ga_results)

            # ── Stats ──
            n_total   = len(ga_res_df)
            n_physio  = (ga_res_df["auto_label"] == "Physiotherapy").sum()
            n_review  = (ga_res_df["needs_review"] == True).sum()
            n_none    = (ga_res_df["auto_label"] == "No Advice").sum()
            n_missed  = (ga_res_df["potential_miss_hints"] != "").sum()
            auto_rate = round((n_physio + n_none) / n_total * 100, 1)

            st.markdown("### 📊 Gap Analysis Results")

            s1, s2, s3, s4, s5 = st.columns(5)
            for col, num, lbl, color in [
                (s1, n_total,  "Total Rows",     "#1a365d"),
                (s2, n_physio, "Auto Physio 🟢",  "#16a34a"),
                (s3, n_review, "Needs Review 🟡", "#f59e0b"),
                (s4, n_none,   "No Advice 🔴",    "#94a3b8"),
                (s5, n_missed, "Possible Misses ⚠️","#dc2626"),
            ]:
                col.markdown(f"""
                <div class="gap-stat">
                    <div class="gap-num" style="color:{color}">{num}</div>
                    <div class="gap-label">{lbl}</div>
                </div>""", unsafe_allow_html=True)

            st.markdown(f"**Auto-label rate: {auto_rate}%** — only **{n_review} rows** need your manual review")

            # ── Keyword frequency chart ──
            all_kws = []
            for kws in ga_res_df["keywords_found"].dropna():
                all_kws.extend([k.strip() for k in kws.split("|") if k.strip()])

            if all_kws:
                st.markdown("### 📈 Keyword Frequency (how often each keyword fires)")
                kw_freq = Counter(all_kws).most_common(20)
                kw_df   = pd.DataFrame(kw_freq, columns=["Keyword", "Count"])
                st.bar_chart(kw_df.set_index("Keyword"))

            # ── Potential missed cases ──
            missed_df = ga_res_df[ga_res_df["potential_miss_hints"] != ""]
            if not missed_df.empty:
                st.markdown(f"### ⚠️ Potential Missed Cases ({len(missed_df)} rows)")
                st.caption("These rows were labeled 'No Advice' but contain physio-adjacent words. Review these first.")
                st.dataframe(
                    missed_df[["patient_id","raw_text","potential_miss_hints"]].head(20),
                    use_container_width=True, hide_index=True
                )

            # ── Needs review rows preview ──
            review_df_preview = ga_res_df[ga_res_df["needs_review"] == True]
            if not review_df_preview.empty:
                st.markdown(f"### 🟡 Rows That Need Your Review ({len(review_df_preview)})")
                st.caption("These will appear in the 'Needs Review' sheet in the downloaded Excel. Fill the MANUAL_LABEL column.")
                st.dataframe(
                    review_df_preview[["patient_id","raw_text","confidence","keywords_found"]].head(10),
                    use_container_width=True, hide_index=True
                )

            # ── Build Excel output ──
            st.markdown("### 💾 Download Gap Analysis Output")
            st.markdown("""
            <div class="step-box">
                <div class="step-num">Your next action</div>
                <div style="margin-top:6px;font-size:0.9rem">
                Download the Excel below → open it → go to the <strong>Needs Review</strong> sheet →
                fill the <code>MANUAL_LABEL</code> column → send it back for ML training.
                </div>
            </div>""", unsafe_allow_html=True)

            buf = io.BytesIO()
            with pd.ExcelWriter(buf, engine="openpyxl") as writer:

                # Sheet 1 — All Results
                all_cols = ["patient_id","raw_text","auto_label","confidence","keywords_found","needs_review","potential_miss_hints"]
                ga_res_df[all_cols].to_excel(writer, index=False, sheet_name="All Results")

                # Sheet 2 — Needs Review (human fills this)
                review_out = ga_res_df[ga_res_df["needs_review"] == True][
                    ["patient_id","raw_text","confidence","keywords_found"]
                ].copy()
                review_out.insert(2, "MANUAL_LABEL", "")
                review_out["NOTES"] = ""
                review_out.to_excel(writer, index=False, sheet_name="Needs Review")

                # Sheet 3 — Confirmed Labels (ML ready)
                confirmed_out = ga_res_df[ga_res_df["needs_review"] == False][
                    ["patient_id","raw_text","auto_label","confidence","keywords_found"]
                ].copy()
                confirmed_out.to_excel(writer, index=False, sheet_name="Confirmed Labels")

                # Sheet 4 — Potential Misses
                if not missed_df.empty:
                    missed_df[["patient_id","raw_text","potential_miss_hints"]].to_excel(
                        writer, index=False, sheet_name="Potential Misses"
                    )

                # Color coding on Sheet 1
                from openpyxl.styles import PatternFill
                ws     = writer.sheets["All Results"]
                green  = PatternFill("solid", fgColor="C6EFCE")
                yellow = PatternFill("solid", fgColor="FFEB9C")
                red    = PatternFill("solid", fgColor="FFC7CE")

                for row_idx, (_, row) in enumerate(ga_res_df[all_cols].iterrows(), start=2):
                    fill = (green  if row["auto_label"] == "Physiotherapy" else
                            yellow if row["auto_label"] == "NEEDS REVIEW"  else red)
                    for col_idx in range(1, len(all_cols)+1):
                        ws.cell(row=row_idx, column=col_idx).fill = fill

                # Auto width all sheets
                for sheet_name in writer.sheets:
                    ws2 = writer.sheets[sheet_name]
                    for col in ws2.columns:
                        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
                        ws2.column_dimensions[col[0].column_letter].width = min(max_len + 4, 60)

            ts = datetime.now().strftime("%Y%m%d_%H%M")
            excel_bytes = buf.getvalue()

            # Also build JSON lines for ML training
            confirmed_rows = ga_res_df[ga_res_df["needs_review"] == False]
            jsonl_lines = []
            for _, row in confirmed_rows.iterrows():
                jsonl_lines.append(json.dumps({
                    "text":       row["raw_text"],
                    "label":      row["auto_label"],
                    "label_int":  1 if row["auto_label"] == "Physiotherapy" else 0,
                    "confidence": row["confidence"],
                    "keywords":   row["keywords_found"],
                }))
            jsonl_bytes = "\n".join(jsonl_lines).encode("utf-8")

            c1, c2 = st.columns(2)
            with c1:
                st.download_button(
                    "⬇️ Download Gap Analysis Excel",
                    excel_bytes,
                    f"gap_analysis_{ts}.xlsx",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    type="primary"
                )
                st.caption("Open this → fill 'Needs Review' sheet → send back for ML training")
            with c2:
                st.download_button(
                    "⬇️ Download ML Training Data (.jsonl)",
                    jsonl_bytes,
                    f"training_data_{ts}.jsonl",
                    "application/json",
                    use_container_width=True
                )
                st.caption("Auto-confirmed labels — ready for Phase 2 ML model training")

            # ── What to do next ──
            st.markdown("---")
            st.markdown("### 🗺️ What Happens Next")
            c1, c2, c3 = st.columns(3)
            with c1:
                st.markdown("""
                <div class="step-box">
                    <div class="step-num">Step A — You do this</div>
                    <div style="margin-top:6px;font-size:0.88rem">
                    Open the Excel → go to <strong>Needs Review</strong> sheet →
                    fill <code>MANUAL_LABEL</code> as <em>Physiotherapy</em> or <em>No Advice</em> →
                    note new keywords in <code>NOTES</code> column
                    </div>
                </div>""", unsafe_allow_html=True)
            with c2:
                st.markdown("""
                <div class="step-box">
                    <div class="step-num">Step B — Send back</div>
                    <div style="margin-top:6px;font-size:0.88rem">
                    Share the filled Excel with your developer.
                    New keywords get added to the dictionary.
                    Model re-runs with improved accuracy.
                    </div>
                </div>""", unsafe_allow_html=True)
            with c3:
                st.markdown("""
                <div class="step-box">
                    <div class="step-num">Step C — Phase 2 ML</div>
                    <div style="margin-top:6px;font-size:0.88rem">
                    Once 500+ rows are labeled, we train a real
                    spaCy NER model that understands context —
                    not just keywords. Much higher accuracy.
                    </div>
                </div>""", unsafe_allow_html=True)
